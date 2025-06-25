import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure the page
st.set_page_config(
    page_title="RATP - ODTGEN Barcode Component Verifier",
    page_icon="📦",
    layout="wide"
)

# Initialize session state
if 'uploaded_df' not in st.session_state:
    st.session_state.uploaded_df = None
if 'scanned_items' not in st.session_state:
    st.session_state.scanned_items = []
if 'known_components' not in st.session_state:
    st.session_state.known_components = set()
if 'clear_input' not in st.session_state:
    st.session_state.clear_input = False
if 'input_counter' not in st.session_state:
    st.session_state.input_counter = 0
if 'continuous_scan_mode' not in st.session_state:
    st.session_state.continuous_scan_mode = False
if 'auto_focus' not in st.session_state:
    st.session_state.auto_focus = False

@st.cache_data
def load_data(uploaded_file):
    """Load and cache the uploaded CSV or Excel data"""
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith('.xlsx'):
            df = pd.read_excel(uploaded_file)
        else:
            st.error("Unsupported file format. Please upload a CSV or Excel file.")
            return None
            
        # Validate required columns
        if 'component_name' not in df.columns or 'component_id' not in df.columns:
            st.error("File must contain 'component_name' and 'component_id' columns")
            return None
        return df
    except Exception as e:
        st.error(f"Error reading file: {str(e)}")
        return None

def add_scanned_item(component_id, component_name=None, status="unknown"):
    """Add a scanned item to the session state"""
    # Check if already scanned to prevent duplicates
    existing_ids = [item['component_id'] for item in st.session_state.scanned_items]
    if component_id not in existing_ids:
        item = {
            'component_id': component_id,
            'component_name': component_name if component_name else "Unknown",
            'status': status,
            'timestamp': datetime.now().strftime("%H:%M:%S")
        }
        st.session_state.scanned_items.append(item)

def check_component_status(component_id):
    """Check if a component ID exists in the uploaded list"""
    if st.session_state.uploaded_df is not None:
        # Convert component_id to string for comparison
        component_id_str = str(component_id)
        uploaded_ids = st.session_state.uploaded_df['component_id'].astype(str)
        
        if component_id_str in uploaded_ids.values:
            # Get component name
            component_row = st.session_state.uploaded_df[uploaded_ids == component_id_str]
            component_name = component_row['component_name'].iloc[0]
            return "known", component_name
    return "unknown", None

def export_unknown_components():
    """Export unknown components as CSV"""
    unknown_items = [item for item in st.session_state.scanned_items if item['status'] == 'unknown']
    if unknown_items:
        df_unknown = pd.DataFrame(unknown_items)
        csv_buffer = io.StringIO()
        df_unknown.to_csv(csv_buffer, index=False)
        return csv_buffer.getvalue()
    return None

def export_complete_results():
    """Export complete results with color coding in Excel format"""
    if st.session_state.uploaded_df is None:
        return None
    
    try:
        # Create a copy of the original dataframe
        df_export = st.session_state.uploaded_df.copy()
        
        # Add scan status column
        scanned_ids = {item['component_id']: item['status'] for item in st.session_state.scanned_items}
        df_export['scan_status'] = df_export['component_id'].astype(str).map(
            lambda x: 'Scanned' if x in scanned_ids else 'Not Scanned'
        )
        
        # Create DataFrame for unknown components (not in original CSV)
        unknown_items = [item for item in st.session_state.scanned_items if item['status'] == 'unknown']
        df_unknown = pd.DataFrame(unknown_items) if unknown_items else pd.DataFrame(columns=['component_id', 'component_name', 'status', 'timestamp'])
        
        # Create Excel workbook
        wb = Workbook()
        
        # Sheet 1: Original components with scan status
        ws1 = wb.active
        ws1.title = "Components Status"
        
        # Add headers
        headers = ['Component Name', 'Component ID', 'Scan Status']
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
            ws1.cell(row=1, column=col).font = Font(bold=True)
        
        # Add data and color coding
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # Light red
        
        for idx, row in df_export.iterrows():
            row_num = idx + 2
            ws1.cell(row=row_num, column=1, value=row['component_name'])
            ws1.cell(row=row_num, column=2, value=str(row['component_id']))
            ws1.cell(row=row_num, column=3, value=row['scan_status'])
            
            # Apply color coding
            if row['scan_status'] == 'Scanned':
                for col in range(1, 4):
                    ws1.cell(row=row_num, column=col).fill = green_fill
            else:
                for col in range(1, 4):
                    ws1.cell(row=row_num, column=col).fill = red_fill
        
        # Sheet 2: Unknown components (not in original CSV)
        if not df_unknown.empty:
            ws2 = wb.create_sheet(title="Unknown Components")
            
            # Add headers for unknown components
            unknown_headers = ['Component ID', 'Time Scanned']
            for col, header in enumerate(unknown_headers, 1):
                ws2.cell(row=1, column=col, value=header)
                ws2.cell(row=1, column=col).font = Font(bold=True)
            
            # Add unknown components data
            orange_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
            
            for idx, row in df_unknown.iterrows():
                row_num = idx + 2
                ws2.cell(row=row_num, column=1, value=str(row['component_id']))
                ws2.cell(row=row_num, column=2, value=row['timestamp'])
                
                # Apply orange color for unknown components
                for col in range(1, 3):
                    ws2.cell(row=row_num, column=col).fill = orange_fill
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error creating Excel export: {str(e)}")
        return None

# Main app layout
st.title("📦 RATP - ODTGEN Barcode scanner")
st.markdown("Upload a CSV or Excel file with component data and scan barcodes to verify components")

# Sidebar
with st.sidebar:
    st.header("Controls")
    
    # File uploader - updated to accept xlsx files
    uploaded_file = st.file_uploader(
        "Upload file",
        type=['csv', 'xlsx'],
        help="File must contain 'component_name' and 'component_id' columns"
    )
    
    if uploaded_file is not None:
        df = load_data(uploaded_file)
        if df is not None:
            st.session_state.uploaded_df = df
            # Create set of known component IDs for faster lookup
            st.session_state.known_components = set(df['component_id'].astype(str))
            st.success(f"✅ Loaded {len(df)} components")
            
            # Display sample of uploaded data
            st.subheader("Sample Data")
            st.dataframe(df.head(), use_container_width=True)
    
    st.divider()
    
    # Counters
    if st.session_state.scanned_items:
        known_count = len([item for item in st.session_state.scanned_items if item['status'] == 'known'])
        unknown_count = len([item for item in st.session_state.scanned_items if item['status'] == 'unknown'])
        
        st.metric("Known Components", known_count, delta=None)
        st.metric("Unknown Components", unknown_count, delta=None)
        st.metric("Total Scanned", len(st.session_state.scanned_items), delta=None)
    
    st.divider()
    
    # Action buttons
    if st.button("🗑️ Clear All Scans", use_container_width=True):
        st.session_state.scanned_items = []
        st.rerun()
    
    # Export unknown components
    if st.session_state.scanned_items:
        unknown_csv = export_unknown_components()
        if unknown_csv:
            st.download_button(
                "📥 Export Unknown Components (CSV)",
                data=unknown_csv,
                file_name=f"unknown_components_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    # Export complete results to Excel
    if st.session_state.uploaded_df is not None:
        excel_data = export_complete_results()
        if excel_data:
            st.download_button(
                "📊 Export Complete Results (Excel)",
                data=excel_data,
                file_name=f"scan_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Export with color coding: Green=Scanned, Red=Not Scanned, Orange=Unknown"
            )

# Main content area
col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("🔍 Barcode Scanner")
    
    # Continuous scan mode toggle
    continuous_mode = st.checkbox(
        "🔄 Continuous Scan Mode",
        value=st.session_state.continuous_scan_mode,
        help="Enable for hands-free continuous barcode scanning"
    )
    st.session_state.continuous_scan_mode = continuous_mode
    
    # Create input key that changes to force re-render and auto-focus
    if st.session_state.clear_input:
        st.session_state.clear_input = False
        st.session_state.input_counter += 1
    
    # Use empty value and force focus with a unique key
    input_key = f"barcode_input_{st.session_state.input_counter}"
    
    # In continuous mode, automatically focus and clear after scan
    if st.session_state.continuous_scan_mode:
        # Inject CSS and JavaScript to auto-focus the input
        st.markdown(f"""
        <style>
        /* Hide the input label for cleaner look in continuous mode */
        .stTextInput > label {{
            font-size: 14px;
            font-weight: 600;
        }}
        </style>
        <script>
        // Auto-focus function
        function autoFocusInput() {{
            // Multiple selectors to find the input field reliably
            const selectors = [
                'input[data-baseweb="input"]',
                'input[type="text"]',
                'input[data-testid*="textInput"]'
            ];
            
            let inputField = null;
            for (const selector of selectors) {{
                const elements = window.parent.document.querySelectorAll(selector);
                for (const element of elements) {{
                    if (element.placeholder && element.placeholder.includes('component ID')) {{
                        inputField = element;
                        break;
                    }}
                }}
                if (inputField) break;
            }}
            
            if (inputField) {{
                inputField.focus();
                inputField.select();
                
                // Add event listener for immediate refocus after Enter
                inputField.addEventListener('keydown', function(e) {{
                    if (e.key === 'Enter') {{
                        setTimeout(() => {{
                            inputField.focus();
                            inputField.select();
                        }}, 100);
                    }}
                }});
            }}
        }}
        
        // Execute multiple times to ensure it works
        setTimeout(autoFocusInput, 100);
        setTimeout(autoFocusInput, 300);
        setTimeout(autoFocusInput, 500);
        setTimeout(autoFocusInput, 1000);
        
        // Observer for DOM changes
        const observer = new MutationObserver(autoFocusInput);
        observer.observe(window.parent.document.body, {{ childList: true, subtree: true }});
        </script>
        """, unsafe_allow_html=True)
    
    # Barcode input
    barcode_input = st.text_input(
        "Scan or enter component ID:",
        value="",
        placeholder="Enter component ID and press Enter...",
        key=input_key
    )
    
    # Auto-scan when text is entered
    if barcode_input and barcode_input.strip():
        if st.session_state.uploaded_df is not None:
            # Check if this component hasn't been scanned yet
            existing_ids = [item['component_id'] for item in st.session_state.scanned_items]
            if barcode_input.strip() not in existing_ids:
                status, component_name = check_component_status(barcode_input.strip())
                add_scanned_item(barcode_input.strip(), component_name, status)
                
                # Show immediate feedback
                if status == "known":
                    st.success(f"✅ Found: {component_name}")
                else:
                    st.warning(f"❌ Unknown component: {barcode_input.strip()}")
                
                # Force refresh for next scan
                st.session_state.clear_input = True
                st.rerun()
            else:
                st.info("This component has already been scanned!")
                st.session_state.clear_input = True
                st.rerun()
        else:
            st.warning("Please upload a CSV or Excel file first!")
    
    # Instructions based on mode
    if st.session_state.continuous_scan_mode:
        st.success("� **Continuous Mode Active**: Input field auto-focuses. Just scan and press Enter!")
        st.caption("💡 Your barcode scanner should automatically enter the code and press Enter. No clicking needed!")
    else:
        st.caption("💡 Tip: Enable Continuous Scan Mode above for hands-free scanning")
    
with col2:
    st.subheader("📋 Scan Results")
    
    if st.session_state.scanned_items:
        # Create DataFrame for display
        df_scanned = pd.DataFrame(st.session_state.scanned_items)
        
        # Add status colors using markdown
        def format_status(status):
            if status == "known":
                return "🟢 Known"
            else:
                return "🔴 Unknown"
        
        df_display = df_scanned.copy()
        df_display['Status'] = df_display['status'].apply(format_status)
        df_display = df_display[['component_id', 'component_name', 'Status', 'timestamp']]
        df_display.columns = ['Component ID', 'Component Name', 'Status', 'Time Scanned']
        
        # Display table
        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Status": st.column_config.TextColumn(
                    "Status",
                    width="small"
                )
            }
        )
        
        # Separate section for unknown components
        unknown_items = [item for item in st.session_state.scanned_items if item['status'] == 'unknown']
        if unknown_items:
            st.subheader("🔴 Unknown Components")
            df_unknown = pd.DataFrame(unknown_items)
            df_unknown_display = df_unknown[['component_id', 'timestamp']]
            df_unknown_display.columns = ['Component ID', 'Time Scanned']
            st.dataframe(df_unknown_display, use_container_width=True, hide_index=True)
    else:
        st.info("No components scanned yet. Upload a CSV or Excel file and start scanning!")

# Footer
st.divider()
st.markdown(
    """
    <div style='text-align: center; color: gray;'>
    📦 RATP - ODTGEN Barcode | By Josselin Perret
    </div>
    """,
    unsafe_allow_html=True
)
