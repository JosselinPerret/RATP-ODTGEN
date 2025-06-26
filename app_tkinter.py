import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import threading
import time

class BarcodeScanner:
    def __init__(self, root):
        self.root = root
        self.root.title("📦 RATP - ODTGEN Barcode Scanner")
        self.root.geometry("1400x800")
        self.root.state('zoomed')  # Maximize window on Windows
        
        # Application data
        self.uploaded_df = None
        self.scanned_items = []
        self.known_components = set()
        self.continuous_scan_mode = True  # Mode scan continu activé par défaut
        
        # Create the interface
        self.create_interface()
        
        # Focus on barcode input by default
        self.root.after(100, self.focus_barcode_input)
        
    def create_interface(self):
        """Create the main interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="📦 RATP - ODTGEN Barcode Scanner", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Left panel - Controls
        self.create_control_panel(main_frame)
        
        # Center panel - Component table
        self.create_component_table_panel(main_frame)
        
        # Right panel - Scan results
        self.create_scan_results_panel(main_frame)
        
        # Status bar
        self.create_status_bar(main_frame)
        
    def create_control_panel(self, parent):
        """Create the left control panel"""
        control_frame = ttk.LabelFrame(parent, text="Controls", padding="10")
        control_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        control_frame.columnconfigure(0, weight=1)
        
        # File operations
        file_frame = ttk.LabelFrame(control_frame, text="File Operations", padding="5")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        self.load_button = ttk.Button(file_frame, text="📁 Load CSV/Excel File", 
                                     command=self.load_file)
        self.load_button.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=2)
        
        self.unload_button = ttk.Button(file_frame, text="🗑️ Unload File", 
                                       command=self.unload_file, state="disabled")
        self.unload_button.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=2)
        
        self.file_info_label = ttk.Label(file_frame, text="No file loaded", 
                                        foreground="gray")
        self.file_info_label.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=2)
        
        # Barcode scanner
        scanner_frame = ttk.LabelFrame(control_frame, text="Barcode Scanner", padding="5")
        scanner_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        scanner_frame.columnconfigure(0, weight=1)
        
        ttk.Label(scanner_frame, text="Scan or enter component ID:").grid(row=0, column=0, sticky=tk.W)
        
        self.barcode_var = tk.StringVar()
        self.barcode_entry = ttk.Entry(scanner_frame, textvariable=self.barcode_var, 
                                      font=("Arial", 12))
        self.barcode_entry.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=2)
        self.barcode_entry.bind('<Return>', self.on_barcode_enter)
        self.barcode_entry.bind('<FocusOut>', self.refocus_barcode_input)
        
        # Continuous scan mode indicator
        self.scan_mode_label = ttk.Label(scanner_frame, text="🔄 Continuous scan mode - Ready", 
                                        foreground="green")
        self.scan_mode_label.grid(row=2, column=0, sticky=tk.W, pady=2)
        
        # Continuous scan mode toggle button
        self.continuous_mode_button = ttk.Button(scanner_frame, text="🔄 Continuous Mode: ON", 
                                                command=self.toggle_continuous_mode)
        self.continuous_mode_button.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=2)
        
        # Statistics
        stats_frame = ttk.LabelFrame(control_frame, text="Statistics", padding="5")
        stats_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.stats_labels = {}
        stats_info = [
            ("total_components", "Total Components: 0"),
            ("scanned_count", "Scanned: 0"),
            ("known_count", "Known: 0"),
            ("unknown_count", "Unknown: 0")
        ]
        
        for i, (key, text) in enumerate(stats_info):
            label = ttk.Label(stats_frame, text=text)
            label.grid(row=i, column=0, sticky=tk.W, pady=1)
            self.stats_labels[key] = label
        
        # Action buttons
        action_frame = ttk.LabelFrame(control_frame, text="Actions", padding="5")
        action_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        action_frame.columnconfigure(0, weight=1)
        
        self.clear_button = ttk.Button(action_frame, text="🗑️ Clear All Scans", 
                                      command=self.clear_scans)
        self.clear_button.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=2)
        
        self.export_unknown_button = ttk.Button(action_frame, text="📥 Export Unknown (CSV)", 
                                               command=self.export_unknown_csv)
        self.export_unknown_button.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=2)
        
        self.export_complete_button = ttk.Button(action_frame, text="📊 Export Complete (Excel)", 
                                                command=self.export_complete_excel)
        self.export_complete_button.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=2)
        
    def create_component_table_panel(self, parent):
        """Create the center panel with component table"""
        table_frame = ttk.LabelFrame(parent, text="Components Table (Click to mark/unmark as scanned)", padding="5")
        table_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)
        
        # Search frame
        search_frame = ttk.Frame(table_frame)
        search_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        search_frame.columnconfigure(1, weight=1)
        
        ttk.Label(search_frame, text="Search:").grid(row=0, column=0, padx=(0, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.grid(row=0, column=1, sticky=(tk.W, tk.E))
        self.search_var.trace('w', self.filter_components)
        
        # Component table with scrollbars
        table_container = ttk.Frame(table_frame)
        table_container.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        table_container.columnconfigure(0, weight=1)
        table_container.rowconfigure(0, weight=1)
        
        # Create treeview for components
        columns = ("Status", "Component ID", "Component Name")
        self.component_tree = ttk.Treeview(table_container, columns=columns, show="headings", height=15)
        
        # Configure columns
        self.component_tree.heading("Status", text="Status")
        self.component_tree.heading("Component ID", text="Component ID")
        self.component_tree.heading("Component Name", text="Component Name")
        
        self.component_tree.column("Status", width=80, anchor="center")
        self.component_tree.column("Component ID", width=150)
        self.component_tree.column("Component Name", width=300)
        
        # Scrollbars for component table
        component_v_scroll = ttk.Scrollbar(table_container, orient="vertical", command=self.component_tree.yview)
        component_h_scroll = ttk.Scrollbar(table_container, orient="horizontal", command=self.component_tree.xview)
        self.component_tree.configure(yscrollcommand=component_v_scroll.set, xscrollcommand=component_h_scroll.set)
        
        # Grid the component table and scrollbars
        self.component_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        component_v_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        component_h_scroll.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Bind click event for manual marking
        self.component_tree.bind("<Button-1>", self.on_component_click)
        
    def create_scan_results_panel(self, parent):
        """Create the right panel with scan results"""
        results_frame = ttk.LabelFrame(parent, text="Scan Results", padding="5")
        results_frame.grid(row=1, column=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # Scan results table
        results_container = ttk.Frame(results_frame)
        results_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        results_container.columnconfigure(0, weight=1)
        results_container.rowconfigure(0, weight=1)
        
        # Create treeview for scan results
        scan_columns = ("Time", "Component ID", "Component Name", "Status")
        self.scan_tree = ttk.Treeview(results_container, columns=scan_columns, show="headings", height=20)
        
        # Configure columns
        self.scan_tree.heading("Time", text="Time")
        self.scan_tree.heading("Component ID", text="Component ID")
        self.scan_tree.heading("Component Name", text="Component Name")
        self.scan_tree.heading("Status", text="Status")
        
        self.scan_tree.column("Time", width=80)
        self.scan_tree.column("Component ID", width=120)
        self.scan_tree.column("Component Name", width=200)
        self.scan_tree.column("Status", width=80, anchor="center")
        
        # Scrollbars for scan results
        scan_v_scroll = ttk.Scrollbar(results_container, orient="vertical", command=self.scan_tree.yview)
        scan_h_scroll = ttk.Scrollbar(results_container, orient="horizontal", command=self.scan_tree.xview)
        self.scan_tree.configure(yscrollcommand=scan_v_scroll.set, xscrollcommand=scan_h_scroll.set)
        
        # Grid the scan results table and scrollbars
        self.scan_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scan_v_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        scan_h_scroll.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
    def create_status_bar(self, parent):
        """Create the status bar"""
        status_frame = ttk.Frame(parent)
        status_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
        status_frame.columnconfigure(1, weight=1)
        
        self.status_label = ttk.Label(status_frame, text="Ready", relief="sunken")
        self.status_label.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Author info
        author_label = ttk.Label(status_frame, text="📦 RATP - ODTGEN Barcode | By Josselin Perret", 
                                foreground="gray")
        author_label.grid(row=0, column=2, sticky=tk.E)
        
    def focus_barcode_input(self):
        """Focus on barcode input field"""
        self.barcode_entry.focus_set()
        
    def refocus_barcode_input(self, event=None):
        """Re-focus on barcode input after a short delay (only in continuous mode)"""
        if self.continuous_scan_mode:
            self.root.after(100, self.focus_barcode_input)
        
    def load_file(self):
        """Load CSV or Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select component file",
            filetypes=[
                ("All supported", "*.csv;*.xlsx"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            try:
                # Load data based on file extension
                if file_path.lower().endswith('.csv'):
                    df = pd.read_csv(file_path)
                elif file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xlsm'):
                    df = pd.read_excel(file_path)
                else:
                    messagebox.showerror("Error", "Unsupported file format. Please select a CSV or Excel file.")
                    return
                
                # Handle different file formats
                if 'component_name' not in df.columns or 'component_id' not in df.columns:
                    # Try to map RATP ODTGEN format (Real/*.csv format)
                    if 'Description' in df.columns and 'Code RIMSES' in df.columns:
                        # Map columns to expected format
                        df = df.rename(columns={
                            'Description': 'component_name',
                            'Code RIMSES': 'component_id'
                        })
                    # Try to map other possible formats
                    elif 'Component Name' in df.columns and 'Component ID' in df.columns:
                        df = df.rename(columns={
                            'Component Name': 'component_name',
                            'Component ID': 'component_id'
                        })
                    elif 'Actif' in df.columns and any(col for col in ['component_name', 'Description'] if col in df.columns):
                        # Test.xlsx format mapping - adapt columns as needed
                        potential_name_cols = [col for col in ['component_name', 'Description'] if col in df.columns]
                        if potential_name_cols:
                            name_col = potential_name_cols[0]
                            id_cols = [col for col in ['component_id', 'Code RIMSES'] if col in df.columns]
                            id_col = id_cols[0] if id_cols else None
                            
                            if id_col:
                                df = df.rename(columns={
                                    name_col: 'component_name',
                                    id_col: 'component_id'
                                })
                    else:
                        messagebox.showerror("Error", "Unsupported file format. The file must contain columns that can be mapped to 'component_name' and 'component_id'")
                        return
                
                # Validate that we now have the required columns
                if 'component_name' not in df.columns or 'component_id' not in df.columns:
                    messagebox.showerror("Error", "Could not identify required columns in the file")
                    return
                
                # Preprocess data to ensure compatibility
                df = self.preprocess_data(df)
                
                # Process and clean the data
                df = self.preprocess_data(df)
                
                # Store data
                self.uploaded_df = df
                self.known_components = set(df['component_id'].astype(str))
                
                # Update file info
                filename = file_path.split('/')[-1].split('\\')[-1]  # Handle both Unix and Windows paths
                file_type = "RATP Format" if "Code RIMSES" in df.columns or "Actif" in df.columns else "Standard Format"
                self.file_info_label.config(
                    text=f"✅ Loaded: {filename} ({len(df)} components) - {file_type}", 
                    foreground="green")
                
                # Enable unload button
                self.unload_button.config(state="normal")
                
                # Populate component table
                self.populate_component_table()
                
                # Update statistics
                self.update_statistics()
                
                # Update status
                self.status_label.config(text=f"Loaded {len(df)} components from {filename}")
                
                # Focus back to barcode input
                self.focus_barcode_input()
                
            except Exception as e:
                messagebox.showerror("Error", f"Error loading file: {str(e)}")
                
    def unload_file(self):
        """Unload the currently loaded file"""
        if messagebox.askyesno("Confirm", "Unload the current file? This will clear all data and scans."):
            # Clear all data
            self.uploaded_df = None
            self.known_components = set()
            self.scanned_items = []
            
            # Update UI
            self.file_info_label.config(text="No file loaded", foreground="gray")
            self.unload_button.config(state="disabled")
            
            # Clear tables
            for item in self.component_tree.get_children():
                self.component_tree.delete(item)
            for item in self.scan_tree.get_children():
                self.scan_tree.delete(item)
            
            # Update statistics
            self.update_statistics()
            
            # Update status
            self.status_label.config(text="File unloaded successfully")
            
            # Focus back to barcode input
            self.focus_barcode_input()
            
    def toggle_continuous_mode(self):
        """Toggle continuous scan mode on/off"""
        self.continuous_scan_mode = not self.continuous_scan_mode
        
        if self.continuous_scan_mode:
            self.continuous_mode_button.config(text="🔄 Continuous Mode: ON")
            self.scan_mode_label.config(text="🔄 Continuous scan mode - Ready", foreground="green")
            # Re-bind events for continuous mode
            self.barcode_entry.bind('<FocusOut>', self.refocus_barcode_input)
        else:
            self.continuous_mode_button.config(text="⏸️ Continuous Mode: OFF")
            self.scan_mode_label.config(text="⏸️ Manual scan mode - Press Enter to scan", foreground="orange")
            # Unbind auto-refocus for manual mode
            self.barcode_entry.unbind('<FocusOut>')
        
        self.status_label.config(text=f"Scan mode: {'Continuous' if self.continuous_scan_mode else 'Manual'}")

    # ...existing code...
    def populate_component_table(self):
        """Populate the component table with loaded data"""
        # Clear existing items
        for item in self.component_tree.get_children():
            self.component_tree.delete(item)
            
        if self.uploaded_df is not None:
            scanned_ids = {item['component_id']: item['status'] for item in self.scanned_items}
            
            for _, row in self.uploaded_df.iterrows():
                component_id = str(row['component_id'])
                component_name = str(row['component_name'])
                
                # Determine status
                if component_id in scanned_ids:
                    status = "✅ Scanned"
                    tags = ("scanned",)
                else:
                    status = "⏳ Pending"
                    tags = ("pending",)
                
                # Insert into tree
                item_id = self.component_tree.insert("", "end", 
                                                    values=(status, component_id, component_name),
                                                    tags=tags)
            
            # Configure tags for color coding
            self.component_tree.tag_configure("scanned", background="#90EE90")  # Light green
            self.component_tree.tag_configure("pending", background="#FFE4B5")   # Light orange
            
    def filter_components(self, *args):
        """Filter components based on search term"""
        if self.uploaded_df is None:
            return
            
        search_term = self.search_var.get().lower()
        
        # Clear existing items
        for item in self.component_tree.get_children():
            self.component_tree.delete(item)
            
        scanned_ids = {item['component_id']: item['status'] for item in self.scanned_items}
        
        for _, row in self.uploaded_df.iterrows():
            component_id = str(row['component_id'])
            component_name = str(row['component_name'])
            
            # Filter based on search term
            if (search_term in component_id.lower() or 
                search_term in component_name.lower() or 
                search_term == ""):
                
                # Determine status
                if component_id in scanned_ids:
                    status = "✅ Scanned"
                    tags = ("scanned",)
                else:
                    status = "⏳ Pending"
                    tags = ("pending",)
                
                # Insert into tree
                self.component_tree.insert("", "end", 
                                          values=(status, component_id, component_name),
                                          tags=tags)
        
        # Configure tags for color coding
        self.component_tree.tag_configure("scanned", background="#90EE90")  # Light green
        self.component_tree.tag_configure("pending", background="#FFE4B5")   # Light orange
        
    def on_component_click(self, event):
        """Handle click on component table for manual marking/unmarking"""
        item = self.component_tree.selection()[0] if self.component_tree.selection() else None
        if item:
            values = self.component_tree.item(item, "values")
            status, component_id, component_name = values
            
            # Check if already scanned
            existing_ids = [item['component_id'] for item in self.scanned_items]
            if component_id not in existing_ids:
                # Mark as manually scanned
                self.add_scanned_item(component_id, component_name, "known", method="manual")
                
                # Show feedback
                self.status_label.config(text=f"Manually marked: {component_name}")
                
                # Provide visual feedback
                self.scan_mode_label.config(text="✅ Component marked manually", foreground="blue")
                self.root.after(2000, lambda: self.scan_mode_label.config(
                    text="🔄 Continuous scan mode - Ready", foreground="green"))
            else:
                # Component already scanned - ask if user wants to unmark it
                if messagebox.askyesno("Unmark Component", 
                                     f"Component '{component_name}' is already scanned.\n\nDo you want to unmark it and put it back to standby?"):
                    # Remove from scanned items
                    self.scanned_items = [item for item in self.scanned_items if item['component_id'] != component_id]
                    
                    # Show feedback
                    self.status_label.config(text=f"Unmarked: {component_name}")
                    
                    # Provide visual feedback
                    self.scan_mode_label.config(text="🔄 Component unmarked - Back to standby", foreground="orange")
                    self.root.after(2000, lambda: self.scan_mode_label.config(
                        text="🔄 Continuous scan mode - Ready", foreground="green"))
            
            # Update displays after any change
            self.populate_component_table()
            self.populate_scan_results()
            self.update_statistics()
            
    def on_barcode_enter(self, event):
        """Handle barcode entry"""
        barcode = self.barcode_var.get().strip()
        if barcode:
            self.process_barcode(barcode)
            # Clear input for next scan
            self.barcode_var.set("")
            
            # Keep focus on input only in continuous mode
            if self.continuous_scan_mode:
                self.barcode_entry.focus_set()
            else:
                # In manual mode, don't auto-refocus
                pass
            
    def process_barcode(self, component_id):
        """Process a scanned barcode"""
        if self.uploaded_df is None:
            messagebox.showwarning("Warning", "Please load a component file first!")
            return
        
        # Check if already scanned
        existing_ids = [item['component_id'] for item in self.scanned_items]
        if component_id in existing_ids:
            self.scan_mode_label.config(text="⚠️ Already scanned!", foreground="orange")
            self.root.after(2000, lambda: self.scan_mode_label.config(
                text="🔄 Continuous scan mode - Ready", foreground="green"))
            return
        
        # Check component status
        status, component_name = self.check_component_status(component_id)
        
        # Add to scanned items
        self.add_scanned_item(component_id, component_name, status)
        
        # Update displays
        self.populate_component_table()
        self.populate_scan_results()
        self.update_statistics()
        
        # Show feedback
        if status == "known":
            self.scan_mode_label.config(text=f"✅ Found: {component_name}", foreground="green")
            self.status_label.config(text=f"Scanned: {component_name}")
        else:
            self.scan_mode_label.config(text=f"❌ Unknown: {component_id}", foreground="red")
            self.status_label.config(text=f"Unknown component: {component_id}")
        
        # Reset scan mode label after 2 seconds
        self.root.after(2000, lambda: self.scan_mode_label.config(
            text="🔄 Continuous scan mode - Ready", foreground="green"))
        
    def check_component_status(self, component_id):
        """Check if a component ID exists in the uploaded list"""
        if self.uploaded_df is not None:
            component_id_str = str(component_id)
            uploaded_ids = self.uploaded_df['component_id'].astype(str)
            
            # Handle the case when Code RIMSES format is used (with ! prefix) or not
            if component_id_str in uploaded_ids.values:
                component_row = self.uploaded_df[uploaded_ids == component_id_str]
                component_name = component_row['component_name'].iloc[0]
                return "known", component_name
            
            # Try without "!" prefix if present in the input
            elif component_id_str.startswith('!') and component_id_str[1:] in uploaded_ids.values:
                component_row = self.uploaded_df[uploaded_ids == component_id_str[1:]]
                component_name = component_row['component_name'].iloc[0]
                return "known", component_name
            
            # Try with "!" prefix if present in the database but not in the input
            elif any(id.startswith('!') and id[1:] == component_id_str for id in uploaded_ids.values):
                for idx, id_val in enumerate(uploaded_ids.values):
                    if id_val.startswith('!') and id_val[1:] == component_id_str:
                        component_name = self.uploaded_df.iloc[idx]['component_name']
                        return "known", component_name
                        
        return "unknown", None
        
    def add_scanned_item(self, component_id, component_name=None, status="unknown", method="barcode"):
        """Add a scanned item"""
        # Check if already scanned to prevent duplicates
        existing_ids = [item['component_id'] for item in self.scanned_items]
        if component_id not in existing_ids:
            item = {
                'component_id': component_id,
                'component_name': component_name if component_name else "Unknown",
                'status': status,
                'method': method,
                'timestamp': datetime.now().strftime("%H:%M:%S")
            }
            self.scanned_items.append(item)
            
    def populate_scan_results(self):
        """Populate the scan results table"""
        # Clear existing items
        for item in self.scan_tree.get_children():
            self.scan_tree.delete(item)
        
        # Add scan results (newest first)
        for item in reversed(self.scanned_items):
            status_symbol = "✅" if item['status'] == "known" else "❌"
            method_symbol = "👆" if item.get('method') == "manual" else "📱"
            
            # Insert with color coding
            tags = ("known",) if item['status'] == "known" else ("unknown",)
            self.scan_tree.insert("", 0, 
                                 values=(f"{method_symbol} {item['timestamp']}", 
                                        item['component_id'], 
                                        item['component_name'], 
                                        status_symbol),
                                 tags=tags)
        
        # Configure tags for color coding
        self.scan_tree.tag_configure("known", background="#90EE90")    # Light green
        self.scan_tree.tag_configure("unknown", background="#FFB6C1")  # Light red
        
    def update_statistics(self):
        """Update statistics labels"""
        total_components = len(self.uploaded_df) if self.uploaded_df is not None else 0
        scanned_count = len(self.scanned_items)
        known_count = len([item for item in self.scanned_items if item['status'] == 'known'])
        unknown_count = len([item for item in self.scanned_items if item['status'] == 'unknown'])
        
        self.stats_labels['total_components'].config(text=f"Total Components: {total_components}")
        self.stats_labels['scanned_count'].config(text=f"Scanned: {scanned_count}")
        self.stats_labels['known_count'].config(text=f"Known: {known_count}")
        self.stats_labels['unknown_count'].config(text=f"Unknown: {unknown_count}")
        
    def clear_scans(self):
        """Clear all scanned items"""
        if messagebox.askyesno("Confirm", "Clear all scanned items?"):
            self.scanned_items = []
            self.populate_component_table()
            self.populate_scan_results()
            self.update_statistics()
            self.status_label.config(text="All scans cleared")
            
    def export_unknown_csv(self):
        """Export unknown components as CSV"""
        unknown_items = [item for item in self.scanned_items if item['status'] == 'unknown']
        if not unknown_items:
            messagebox.showinfo("Info", "No unknown components to export")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Save unknown components",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                df_unknown = pd.DataFrame(unknown_items)
                df_unknown.to_csv(file_path, index=False)
                messagebox.showinfo("Success", f"Unknown components exported to {file_path}")
                self.status_label.config(text=f"Exported {len(unknown_items)} unknown components")
            except Exception as e:
                messagebox.showerror("Error", f"Error exporting file: {str(e)}")
                
    def export_complete_excel(self):
        """Export complete results with color coding in Excel format"""
        if self.uploaded_df is None:
            messagebox.showwarning("Warning", "No data to export")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Save complete results",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                # Create a copy of the original dataframe
                df_export = self.uploaded_df.copy()
                
                # Add scan status column
                scanned_ids = {item['component_id']: item['status'] for item in self.scanned_items}
                df_export['scan_status'] = df_export['component_id'].astype(str).map(
                    lambda x: 'Scanned' if x in scanned_ids else 'Not Scanned'
                )
                
                # Create DataFrame for unknown components
                unknown_items = [item for item in self.scanned_items if item['status'] == 'unknown']
                df_unknown = pd.DataFrame(unknown_items) if unknown_items else pd.DataFrame(
                    columns=['component_id', 'component_name', 'status', 'timestamp'])
                
                # Create Excel workbook
                wb = Workbook()
                
                # Sheet 1: Original components with scan status
                ws1 = wb.active
                ws1.title = "Components Status"
                
                # Add headers
                headers = ['Component Name', 'Component ID', 'Scan Status']
                for col, header in enumerate(headers, 1):
                    ws1.cell(row=1, column=col, value=header)
                    ws1.cell(row=1, column=col).font = Font(bold=True)
                
                # Add data and color coding
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                for idx, row in df_export.iterrows():
                    row_num = idx + 2
                    ws1.cell(row=row_num, column=1, value=row['component_name'])
                    ws1.cell(row=row_num, column=2, value=str(row['component_id']))
                    ws1.cell(row=row_num, column=3, value=row['scan_status'])
                    
                    # Apply color coding
                    if row['scan_status'] == 'Scanned':
                        for col in range(1, 4):
                            ws1.cell(row=row_num, column=col).fill = green_fill
                    else:
                        for col in range(1, 4):
                            ws1.cell(row=row_num, column=col).fill = red_fill
                
                # Sheet 2: Unknown components
                if not df_unknown.empty:
                    ws2 = wb.create_sheet(title="Unknown Components")
                    
                    unknown_headers = ['Component ID', 'Time Scanned', 'Method']
                    for col, header in enumerate(unknown_headers, 1):
                        ws2.cell(row=1, column=col, value=header)
                        ws2.cell(row=1, column=col).font = Font(bold=True)
                    
                    orange_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    
                    for idx, row in df_unknown.iterrows():
                        row_num = idx + 2
                        ws2.cell(row=row_num, column=1, value=str(row['component_id']))
                        ws2.cell(row=row_num, column=2, value=row['timestamp'])
                        ws2.cell(row=row_num, column=3, value=row.get('method', 'barcode'))
                        
                        for col in range(1, 4):
                            ws2.cell(row=row_num, column=col).fill = orange_fill
                
                # Auto-adjust column widths
                for sheet in wb.worksheets:
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Save workbook
                wb.save(file_path)
                
                messagebox.showinfo("Success", f"Complete results exported to {file_path}")
                self.status_label.config(text="Complete results exported successfully")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error exporting file: {str(e)}")
    
    def preprocess_data(self, df):
        """Preprocess the data to ensure compatibility with the application"""
        # Handle various formats and clean up the data
        
        # If we have Code RIMSES column, ensure the ! is handled properly
        if 'Code RIMSES' in df.columns:
            # Rename to component_id if needed
            if 'component_id' not in df.columns:
                df = df.rename(columns={'Code RIMSES': 'component_id'})
        
        # Ensure component_id is string type
        if 'component_id' in df.columns:
            df['component_id'] = df['component_id'].astype(str)
            
            # Remove any spaces or special characters that might cause issues
            df['component_id'] = df['component_id'].str.strip()
            
        # Ensure component_name is string type
        if 'component_name' in df.columns:
            df['component_name'] = df['component_name'].astype(str)
            df['component_name'] = df['component_name'].str.strip()
        
        # Add actif as identifier if it exists and might be useful
        if 'Actif' in df.columns:
            df['actif'] = df['Actif'].astype(str)
            
        return df

def main():
    root = tk.Tk()
    app = BarcodeScanner(root)
    root.mainloop()

if __name__ == "__main__":
    main()
